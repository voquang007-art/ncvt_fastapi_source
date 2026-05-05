from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import unicodedata

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import summary_required
from ..time_utils import vietnam_now
from ..models import (
    DemandBalanceResult,
    DemandRequest,
    DemandRequestItem,
    InventorySnapshot,
    InventorySnapshotItem,
    ROLE_LABELS,
    STATUS_CLOSED,
    STATUS_SUMMARIZED,
    User,
)

router = APIRouter(prefix="/inventory-balance", tags=["inventory_balance"])
templates = Jinja2Templates(directory="app/templates")
templates.env.globals.update(ROLE_LABELS=ROLE_LABELS, NAV_MONTH=date.today().month, NAV_YEAR=date.today().year)


def _norm(value: object) -> str:
    return str(value or "").strip()


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    text = " ".join(text.replace("_", " ").split())
    return text


def _norm_key(value: object) -> str:
    return _norm(value).upper()


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _find_col(headers: list[str], candidates: list[str]) -> int | None:
    normalized_headers = [_norm_header(header) for header in headers]
    normalized_candidates = [_norm_header(candidate) for candidate in candidates]

    for candidate in normalized_candidates:
        if candidate in normalized_headers:
            return normalized_headers.index(candidate)

    return None


def get_active_snapshot(db: Session, month: int, year: int) -> InventorySnapshot | None:
    return (
        db.query(InventorySnapshot)
        .filter(
            InventorySnapshot.month == month,
            InventorySnapshot.year == year,
            InventorySnapshot.is_active == True,
        )
        .order_by(InventorySnapshot.imported_at.desc())
        .first()
    )


def collect_demand_totals(db: Session, month: int, year: int) -> list[dict]:
    reqs = (
        db.query(DemandRequest)
        .filter(
            DemandRequest.month == month,
            DemandRequest.year == year,
            DemandRequest.status.in_([STATUS_SUMMARIZED, STATUS_CLOSED]),
            DemandRequest.is_cancelled == False,
        )
        .all()
    )
    req_ids = [r.id for r in reqs]
    if not req_ids:
        return []

    items = db.query(DemandRequestItem).filter(DemandRequestItem.request_id.in_(req_ids)).all()
    grouped: dict[tuple[str, str, str, str], dict] = {}
    for item in items:
        key = (
            _norm_key(item.material_code) if item.material_code else "",
            _norm(item.material_name),
            _norm(item.unit),
            _norm(item.category_name),
        )
        if key not in grouped:
            grouped[key] = {
                "material_code": item.material_code or "",
                "material_name": item.material_name,
                "unit": item.unit,
                "category_name": item.category_name,
                "total_demand_qty": 0.0,
                "is_new_material": bool(item.is_new_material),
            }
        grouped[key]["total_demand_qty"] += float(item.quantity or 0)
        if item.is_new_material:
            grouped[key]["is_new_material"] = True

    rows = list(grouped.values())
    rows.sort(key=lambda x: (x["category_name"], x["material_name"]))
    return rows


def collect_balance_rows(db: Session, month: int, year: int) -> tuple[InventorySnapshot | None, list[dict]]:
    snapshot = get_active_snapshot(db, month, year)
    demands = collect_demand_totals(db, month, year)

    stock_by_code: dict[str, dict] = {}
    if snapshot:
        for item in snapshot.items:
            code = _norm_key(item.material_code)
            if not code:
                continue

            unit = _norm(item.unit)
            stock_qty = float(item.stock_qty or 0)
            existing = stock_by_code.get(code)

            if not existing:
                stock_by_code[code] = {
                    "unit": unit,
                    "stock_qty": stock_qty,
                    "duplicate_count": 1,
                    "unit_conflict": False,
                }
                continue

            existing["duplicate_count"] += 1
            if _norm_key(existing["unit"]) == _norm_key(unit):
                existing["stock_qty"] += stock_qty
            else:
                existing["unit_conflict"] = True

    rows: list[dict] = []
    for demand in demands:
        material_code = _norm(demand["material_code"])
        code_key = _norm_key(material_code)
        stock_qty = 0.0
        status_note = ""

        if demand.get("is_new_material") or not code_key:
            status_note = "Vật tư ngoài danh mục/chưa có mã, không đối chiếu tồn kho tự động."
        elif not snapshot:
            status_note = "Chưa import tồn kho cho tháng này."
        else:
            stock_info = stock_by_code.get(code_key)
            if not stock_info:
                status_note = "Không có trong file tồn kho tháng này."
            elif stock_info.get("unit_conflict"):
                status_note = "File tồn kho có nhiều dòng cùng mã nhưng khác đơn vị tính, không tự động cộng dồn."
            elif _norm_key(stock_info["unit"]) != _norm_key(demand["unit"]):
                status_note = f"Khác đơn vị tính tồn kho ({stock_info['unit']}) và nhu cầu ({demand['unit']})."
            else:
                stock_qty = float(stock_info["stock_qty"] or 0)
                if stock_info.get("duplicate_count", 1) > 1:
                    status_note = "Tồn kho có nhiều dòng cùng mã, đã cộng dồn theo mã và đơn vị tính."

        total_demand = float(demand["total_demand_qty"] or 0)
        shortage = max(total_demand - stock_qty, 0)
        surplus = max(stock_qty - total_demand, 0)
        rows.append(
            {
                "material_code": material_code or "Ngoài danh mục",
                "material_name": demand["material_name"],
                "unit": demand["unit"],
                "category_name": demand["category_name"],
                "total_demand_qty": total_demand,
                "stock_qty": stock_qty,
                "shortage_qty": shortage,
                "surplus_qty": surplus,
                "suggested_purchase_qty": shortage,
                "status_note": status_note,
                "snapshot_id": snapshot.id if snapshot else None,
            }
        )
    return snapshot, rows


@router.get("", response_class=HTMLResponse)
def balance_page(
    request: Request,
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(summary_required),
):
    snapshot, rows = collect_balance_rows(db, month, year)
    saved_count = db.query(DemandBalanceResult).filter(
        DemandBalanceResult.month == month,
        DemandBalanceResult.year == year,
    ).count()
    return templates.TemplateResponse(
        "inventory_balance.html",
        {
            "request": request,
            "current_user": current_user,
            "month": month,
            "year": year,
            "snapshot": snapshot,
            "rows": rows,
            "saved_count": saved_count,
            "today": date.today(),
        },
    )


@router.post("/import")
async def import_inventory_snapshot(
    month: int = Form(...),
    year: int = Form(...),
    snapshot_date: str = Form(...),
    note: str = Form(""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(summary_required),
):
    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = [_norm_header(cell.value) for cell in ws[1]]

    col_code = _find_col(headers, ["mã vật tư", "ma vat tu", "mã sản phẩm", "ma san pham", "code"])
    col_name = _find_col(headers, ["tên vật tư", "ten vat tu", "danh mục", "danh muc", "name"])
    col_unit = _find_col(headers, ["đơn vị tính", "don vi tinh", "đơn vị tinh", "don vi tinh", "đvt", "dvt", "unit"])
    col_cat = _find_col(headers, ["loại vật tư", "loai vat tu", "loại hh", "loai hh", "category"])
    col_stock = _find_col(headers, ["tồn kho", "ton kho", "số lượng tồn", "so luong ton", "stock", "stock_qty"])
    col_note = _find_col(headers, ["ghi chú", "ghi chu", "note"])

    if None in (col_code, col_name, col_unit, col_stock):
        return RedirectResponse(f"/inventory-balance?month={month}&year={year}&error=invalid_template", status_code=303)

    old_snapshots = db.query(InventorySnapshot).filter(
        InventorySnapshot.month == month,
        InventorySnapshot.year == year,
        InventorySnapshot.is_active == True,
    ).all()
    for old in old_snapshots:
        old.is_active = False

    snapshot = InventorySnapshot(
        month=month,
        year=year,
        snapshot_date=date.fromisoformat(snapshot_date),
        source_file_name=file.filename,
        note=note.strip() or None,
        is_active=True,
        imported_by=current_user.id,
        imported_at=vietnam_now(),
    )
    db.add(snapshot)
    db.flush()

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = _norm(row[col_code])
        name = _norm(row[col_name])
        unit = _norm(row[col_unit])
        stock_qty = _to_float(row[col_stock])
        if not code or not name or not unit:
            continue
        db.add(
            InventorySnapshotItem(
                snapshot_id=snapshot.id,
                material_code=code,
                material_name=name,
                unit=unit,
                category_name=_norm(row[col_cat]) if col_cat is not None else None,
                stock_qty=stock_qty,
                note=_norm(row[col_note]) if col_note is not None else None,
            )
        )

    db.commit()
    return RedirectResponse(f"/inventory-balance?month={month}&year={year}", status_code=303)


@router.post("/save")
def save_balance_result(
    month: int = Form(...),
    year: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(summary_required),
):
    snapshot, rows = collect_balance_rows(db, month, year)
    if not snapshot:
        return RedirectResponse(
            f"/inventory-balance?month={month}&year={year}&error=no_snapshot",
            status_code=303,
        )

    db.query(DemandBalanceResult).filter(
        DemandBalanceResult.month == month,
        DemandBalanceResult.year == year,
    ).delete(synchronize_session=False)

    for row in rows:
        db.add(
            DemandBalanceResult(
                month=month,
                year=year,
                material_code=row["material_code"],
                material_name=row["material_name"],
                unit=row["unit"],
                category_name=row["category_name"],
                total_demand_qty=row["total_demand_qty"],
                stock_qty=row["stock_qty"],
                shortage_qty=row["shortage_qty"],
                surplus_qty=row["surplus_qty"],
                suggested_purchase_qty=row["suggested_purchase_qty"],
                status_note=row["status_note"],
                snapshot_id=snapshot.id if snapshot else None,
                created_by=current_user.id,
            )
        )
    db.commit()
    return RedirectResponse(f"/inventory-balance?month={month}&year={year}", status_code=303)


@router.get("/export.xlsx")
def export_balance_excel(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(summary_required),
):
    snapshot, rows = collect_balance_rows(db, month, year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cân đối tồn kho"
    ws.append([f"BẢNG CÂN ĐỐI NHU CẦU VẬT TƯ THÁNG {month}/{year}"])
    ws.append(["Ngày chốt tồn", snapshot.snapshot_date.strftime("%d/%m/%Y") if snapshot else "Chưa import tồn kho"])
    ws.append([])
    ws.append(["STT", "Mã vật tư", "Tên vật tư", "Loại", "ĐVT", "Tổng nhu cầu", "Tồn kho tháng", "Thiếu", "Dư", "Đề xuất mua", "Ghi chú"])
    for idx, row in enumerate(rows, 1):
        ws.append([
            idx,
            row["material_code"],
            row["material_name"],
            row["category_name"],
            row["unit"],
            row["total_demand_qty"],
            row["stock_qty"],
            row["shortage_qty"],
            row["surplus_qty"],
            row["suggested_purchase_qty"],
            row["status_note"],
        ])

    bio = BytesIO()
    wb.save(bio)
    content = bio.getvalue()
    filename = f"can_doi_nhu_cau_vat_tu_{year}_{month:02d}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
