from datetime import date
from io import BytesIO
import unicodedata

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import login_required
from ..models import Material, ROLE_ADMIN, ROLE_LABELS, ROLE_SUMMARY, User

router = APIRouter(prefix="/materials", tags=["materials"])
templates = Jinja2Templates(directory="app/templates")
templates.env.globals.update(ROLE_LABELS=ROLE_LABELS, NAV_MONTH=date.today().month, NAV_YEAR=date.today().year)


def can_manage_materials(user: User) -> bool:
    return user.role in {ROLE_ADMIN, ROLE_SUMMARY}


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    text = " ".join(text.replace("_", " ").split())
    return text


@router.get("", response_class=HTMLResponse)
def materials_page(
    request: Request,
    q: str = "",
    category: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(login_required),
):
    categories = [
        row[0]
        for row in db.query(Material.category_name)
        .filter(Material.category_name.isnot(None), Material.category_name != "")
        .distinct()
        .order_by(Material.category_name)
        .all()
    ]

    query = db.query(Material)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            (Material.code.like(like))
            | (Material.name.like(like))
            | (Material.category_name.like(like))
        )

    if category:
        query = query.filter(Material.category_name == category)

    materials = query.order_by(Material.category_name, Material.name).all()
    return templates.TemplateResponse(
        "materials.html",
        {
            "request": request,
            "current_user": current_user,
            "materials": materials,
            "q": q,
            "categories": categories,
            "selected_category": category,
            "can_manage": can_manage_materials(current_user),
        },
    )


@router.post("/create")
def create_material(
    code: str = Form(...),
    name: str = Form(...),
    unit: str = Form(...),
    category_name: str = Form(...),
    group_name: str = Form(""),
    specification: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(login_required),
):
    if not can_manage_materials(current_user):
        return RedirectResponse("/materials", status_code=303)
    mat = db.query(Material).filter(Material.code == code.strip()).first()
    if not mat:
        db.add(
            Material(
                code=code.strip(),
                name=name.strip(),
                unit=unit.strip(),
                category_name=category_name.strip(),
                group_name=group_name.strip() or None,
                specification=specification.strip() or None,
                note=note.strip() or None,
            )
        )
    else:
        mat.name = name.strip()
        mat.unit = unit.strip()
        mat.category_name = category_name.strip()
        mat.group_name = group_name.strip() or None
        mat.specification = specification.strip() or None
        mat.note = note.strip() or None
    db.commit()
    return RedirectResponse("/materials", status_code=303)


@router.post("/import")
async def import_materials(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(login_required),
):
    if not can_manage_materials(current_user):
        return RedirectResponse("/materials", status_code=303)

    content = await file.read()
    if not content:
        return RedirectResponse("/materials?error=empty_file", status_code=303)

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        return RedirectResponse("/materials?error=invalid_excel", status_code=303)

    ws = wb.active
    first_row = [cell.value for cell in ws[1]]
    headers = [_norm_header(cell.value) for cell in ws[1]]

    def find_col(names):
        normalized_names = [_norm_header(name) for name in names]
        for name in normalized_names:
            if name in headers:
                return headers.index(name)
        return None

    col_code = find_col(["mã vật tư", "ma vat tu", "mã sản phẩm", "ma san pham", "code"])
    col_name = find_col(["tên vật tư", "ten vat tu", "danh mục", "danh muc", "name"])
    col_unit = find_col(["đơn vị tính", "don vi tinh", "đvt", "dvt", "unit"])
    col_cat = find_col(["loại vật tư", "loai vat tu", "loại hh", "loai hh", "category"])
    col_group = find_col(["nhóm vật tư", "nhom vat tu", "group"])
    col_spec = find_col(["quy cách", "quy cach", "specification"])
    col_note = find_col(["ghi chú", "ghi chu", "note"])

    has_header = None not in (col_code, col_name, col_unit, col_cat)

    if has_header:
        start_row = 2
    else:
        non_empty_first_row = [str(value or "").strip() for value in first_row]
        if len(non_empty_first_row) < 4 or not all(non_empty_first_row[:4]):
            return RedirectResponse("/materials?error=invalid_template", status_code=303)

        col_code = 0
        col_name = 1
        col_unit = 2
        col_cat = 3
        col_group = None
        col_spec = 4 if ws.max_column >= 5 else None
        col_note = 5 if ws.max_column >= 6 else None
        start_row = 1

    imported_count = 0
    skipped_count = 0

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        code = str(row[col_code] or "").strip() if col_code is not None and col_code < len(row) else ""
        name = str(row[col_name] or "").strip() if col_name is not None and col_name < len(row) else ""
        unit = str(row[col_unit] or "").strip() if col_unit is not None and col_unit < len(row) else ""
        cat = str(row[col_cat] or "").strip() if col_cat is not None and col_cat < len(row) else ""

        if not code or not name or not unit or not cat:
            skipped_count += 1
            continue

        mat = db.query(Material).filter(Material.code == code).first()
        if not mat:
            mat = Material(code=code, name=name, unit=unit, category_name=cat)
            db.add(mat)

        mat.name = name
        mat.unit = unit
        mat.category_name = cat
        mat.group_name = (
            str(row[col_group] or "").strip()
            if col_group is not None and col_group < len(row)
            else None
        )
        mat.specification = (
            str(row[col_spec] or "").strip()
            if col_spec is not None and col_spec < len(row)
            else None
        )
        mat.note = (
            str(row[col_note] or "").strip()
            if col_note is not None and col_note < len(row)
            else None
        )
        imported_count += 1

    db.commit()
    return RedirectResponse(
        f"/materials?imported={imported_count}&skipped={skipped_count}",
        status_code=303,
    )