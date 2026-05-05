from __future__ import annotations

from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import DemandRequest


BRANDING_LOGO = Path(__file__).resolve().parent / "static" / "branding" / "logo.png"


def register_vietnamese_font() -> str:
    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/times.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for path in candidates:
        if path.exists():
            try:
                pdfmetrics.registerFont(TTFont("VNFont", str(path)))
                return "VNFont"
            except Exception:
                pass
    return "Helvetica"


def fmt_qty(value: float) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return f"{int(value):,}".replace(",", ".")
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _build_header(font: str, normal: ParagraphStyle, unit_name: str):
    company_style = ParagraphStyle(
        "CompanyStyle",
        parent=normal,
        fontName=font,
        fontSize=11,
        leading=14,
        textColor=colors.red,
        alignment=TA_LEFT,
        spaceAfter=2,
    )
    unit_style = ParagraphStyle(
        "UnitStyle",
        parent=normal,
        fontName=font,
        fontSize=10,
        leading=13,
        textColor=colors.black,
        alignment=TA_LEFT,
    )

    if BRANDING_LOGO.exists():
        left_cell = Image(str(BRANDING_LOGO), width=1.8 * cm, height=1.8 * cm)
    else:
        left_cell = ""

    right_cell = [
        Paragraph("<b>BỆNH VIỆN HÙNG VƯƠNG GIA LAI</b>", company_style),
        Paragraph(f"<b>Đơn vị: {escape(unit_name)}</b>", unit_style),
    ]

    header_table = Table([[left_cell, right_cell]], colWidths=[2.2 * cm, 14.0 * cm])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return header_table


def build_request_pdf(req: DemandRequest) -> bytes:
    font = register_vietnamese_font()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    normal = ParagraphStyle("VNNormal", parent=styles["Normal"], fontName=font, fontSize=10, leading=13)
    center = ParagraphStyle("VNCenter", parent=normal, alignment=TA_CENTER)
    title = ParagraphStyle("VNTitle", parent=center, fontName=font, fontSize=12, leading=15, spaceAfter=4)
    small_center = ParagraphStyle("VNSmallCenter", parent=center, fontSize=9)
    right_style = ParagraphStyle("RightStyle", parent=normal, alignment=TA_CENTER, leftIndent=10 * cm)

    categories = sorted({
        (item.category_name or "").strip()
        for item in req.items
        if (item.category_name or "").strip()
    })
    categories_text = ", ".join(categories) if categories else "vật tư/hàng hóa"
    categories_title = categories_text.upper()

    header_table = _build_header(font, normal, req.unit.name)

    story = []
    story.append(header_table)
    story.append(Spacer(1, 0.25 * cm))
    story.append(Paragraph(f"<b>NHU CẦU SỬ DỤNG {escape(categories_title)}</b>", title))
    story.append(Paragraph(f"<b>Tháng {req.month} năm {req.year}</b>", center))
    story.append(Spacer(1, 0.25 * cm))
    recipient_table = Table(
        [
            [
                Paragraph("<b>Kính gửi:</b>", normal),
                Paragraph(
                    "- Lãnh đạo Công ty<br/>"
                    "- Đơn vị Mua sắm<br/>"
                    "- Phòng Tài chính Kế toán",
                    normal,
                ),
            ]
        ],
        colWidths=[2.4 * cm, 12.8 * cm],
    )
    recipient_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(recipient_table)
    story.append(Spacer(1, 0.25 * cm))
    story.append(
        Paragraph(
            f"Thực hiện nhiệm vụ trong tháng {req.month} năm {req.year}. "
            f"Đơn vị {escape(req.unit.name)} lập nhu cầu sử dụng {escape(categories_text)} "
            f"đề nghị mua sắm như sau:",
            normal,
        )
    )
    story.append(Spacer(1, 0.25 * cm))

    data = [[
        Paragraph("<b>STT</b>", center),
        Paragraph("<b>Danh mục</b>", center),
        Paragraph("<b>Mã sản phẩm</b>", center),
        Paragraph("<b>Đơn vị tính</b>", center),
        Paragraph("<b>Số lượng</b>", center),
        Paragraph("<b>Loại HH</b>", center),
        Paragraph("<b>Quy cách</b>", center),
    ]]
    for idx, item in enumerate(req.items, 1):
        data.append([
            str(idx),
            Paragraph(escape(item.material_name or ""), normal),
            Paragraph(escape(item.material_code or "Ngoài danh mục"), normal),
            Paragraph(escape(item.unit or ""), center),
            fmt_qty(item.quantity),
            Paragraph(escape(item.category_name or ""), normal),
            Paragraph(escape(item.specification or item.note or ""), normal),
        ])

    table = Table(data, colWidths=[1.0 * cm, 4.7 * cm, 3.0 * cm, 1.8 * cm, 1.7 * cm, 3.2 * cm, 3.2 * cm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (4, -1), "CENTER"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.55 * cm))
    story.append(
        Paragraph(
            f"Gia Lai, ngày {req.request_date.day} tháng {req.request_date.month} năm {req.request_date.year}",
            right_style,
        )
    )
    story.append(Spacer(1, 0.2 * cm))
    sign = Table(
        [
            [Paragraph("<b>NGƯỜI LẬP</b>", center), Paragraph("<b>PHỤ TRÁCH KHOA/PHÒNG</b>", center)],
            [Paragraph("(Ký, ghi rõ họ tên)", small_center), Paragraph("(Ký, ghi rõ họ tên)", small_center)],
            ["", ""],
            [Paragraph(escape(req.input_person_name or ""), center), Paragraph(escape(req.unit_head_name or ""), center)],
        ],
        colWidths=[8.5 * cm, 8.5 * cm],
    )
    sign.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOTTOMPADDING", (0, 2), (-1, 4), 24),
            ]
        )
    )
    story.append(sign)

    doc.build(story)
    return buffer.getvalue()


def build_summary_excel(rows: list[dict], detail_rows: list[dict], new_material_rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng hợp vật tư"
    ws.append(["STT", "Mã vật tư", "Tên vật tư", "Loại", "ĐVT", "Tổng số lượng", "Số đơn vị đăng ký"])
    for idx, row in enumerate(rows, 1):
        ws.append([idx, row["material_code"], row["material_name"], row["category_name"], row["unit"], row["total_quantity"], row["unit_count"]])

    ws2 = wb.create_sheet("Chi tiết đơn vị")
    ws2.append(["STT", "Đơn vị", "Mã vật tư", "Tên vật tư", "Loại", "ĐVT", "Số lượng", "Ghi chú"])
    for idx, row in enumerate(detail_rows, 1):
        ws2.append([idx, row["unit_name"], row["material_code"], row["material_name"], row["category_name"], row["unit"], row["quantity"], row["note"]])

    ws3 = wb.create_sheet("Ngoài danh mục")
    ws3.append(["STT", "Đơn vị", "Tên vật tư đề xuất", "Loại", "ĐVT", "Số lượng", "Ghi chú"])
    for idx, row in enumerate(new_material_rows, 1):
        ws3.append([idx, row["unit_name"], row["material_name"], row["category_name"], row["unit"], row["quantity"], row["note"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
